#Binex IT
#Developer: Eric Joh
#Volume and Profit Report Generator
#Nov 2020

import sys
import os
import openpyxl
#from openpyxl import load_workbook, Workbook
#from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk

class Node:
    def __init__(self, month, location, agent, vol, profit):
        self.month = month
        self.location = location
        self.agent = agent
        self.vol = vol
        self.profit = profit

titlefill = openpyxl.styles.PatternFill(fill_type='solid', start_color='fdc16b', end_color='fdc16b')
datafill = openpyxl.styles.PatternFill(fill_type='solid', start_color='fae5b0', end_color='fae5b0')

def title(v1, row, maxmonth, office, tabletype, monthval):
    #tabletype= 1: By agent, 2: By office, 3: By month
    col = 1
    if tabletype == 1:
        v1.cell(row=(row-1), column=col).value = "By Agent"
        v1.cell(row=row, column=col).value = "Office"
    elif tabletype == 2:
        v1.cell(row=(row-1), column=col).value = "By Office"
        v1.cell(row=row, column=col).value = "Agent"
    elif tabletype == 3:
        v1.cell(row=(row-1), column=col).value = "By Month (" + monthval + ")"
        v1.cell(row=row, column=col).value = "Agent"     
    else:
        print("Wrong input function - \'title part1\'")
    
    v1.cell(row=row, column=col).fill = titlefill
    v1.cell(row=row, column=col).font = openpyxl.styles.Font(bold = True)
    v1.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
    
    if tabletype == 1 or tabletype == 2:
        while (col <= maxmonth):
            col += 1
            v1.cell(row=row, column=col).value = col - 1
            v1.cell(row=row, column=col).font = openpyxl.styles.Font(bold = True)
            v1.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
            v1.cell(row=row, column=col).fill = titlefill
        v1.cell(row=row, column=(col+1)).value = "Grand Total"
        v1.cell(row=row, column=(col+1)).font = openpyxl.styles.Font(bold = True)
        v1.cell(row=row, column=(col+1)).alignment = openpyxl.styles.Alignment(horizontal='center')
        v1.cell(row=row, column=(col+1)).fill = titlefill
    elif tabletype == 3:
        while (col <= len(office)):
            col += 1
            v1.cell(row=row, column=col).value = office[col - 2]
            v1.cell(row=row, column=col).font = openpyxl.styles.Font(bold = True)
            v1.cell(row=row, column=col).alignment = openpyxl.styles.Alignment(horizontal='center')
            v1.cell(row=row, column=col).fill = titlefill
        v1.cell(row=row, column=(col+1)).value = "Grand Total"
        v1.cell(row=row, column=(col+1)).font = openpyxl.styles.Font(bold = True)
        v1.cell(row=row, column=(col+1)).alignment = openpyxl.styles.Alignment(horizontal='center')
        v1.cell(row=row, column=(col+1)).fill = titlefill
    else:
        print("Wrong input function - \'title part2\'")
    
    #Merge cell of title
    merge = openpyxl.utils.get_column_letter(1) + str(row-1) + ":" + openpyxl.utils.get_column_letter(col + 1) + str(row-1)
    v1.merge_cells(merge)
    v1.cell(row = (row - 1), column = 1).fill = titlefill
    v1.cell(row = (row - 1), column = 1).font = openpyxl.styles.Font(italic = True, bold = True)
    v1.cell(row = (row - 1), column = 1).alignment = openpyxl.styles.Alignment(horizontal='center')

    row += 1
    return row

def label(v1, row, office, person, tabletype):
    #tabletype= 1: By agent, 2: By office, 3: By month
    i = 0
    temp = row
    if tabletype == 1:
        while row <= (len(office) + (temp - 1)):
            v1.cell(row=row, column=1).value = office[i]
            v1.cell(row=row, column=1).fill = datafill
            i += 1
            row += 1
    elif tabletype == 2 or tabletype == 3:
        while row <= (len(person) + temp - 1):
            v1.cell(row=row, column=1).value = person[i]
            v1.cell(row=row, column=1).fill = datafill
            i += 1
            row += 1
    else:
        print("Wrong input function - \'Lable\'")

def dataset(v1, node, row, office, person, maxmonth, tabletype):
    #tabletype= 1: By agent, 2: By office, 3: By month
    i = 0
    #reference = openpyxl.utils.get_column_letter(2) + str(row) + ":" + openpyxl.utils.get_column_letter(maxmonth + 2)
    #condition = openpyxl.utils.get_column_letter(maxmonth + 2) + str(row) + ":" + openpyxl.utils.get_column_letter(maxmonth + 2)
    if tabletype == 1:
        grandtotal = v1.cell(row=(row + len(office)), column=1)
        #data_main = v1.cell(row=(row + office.index(node[i].location)), column=(node[i].month + 1))
        #data_row = v1.cell(row=(row + office.index(node[i].location)), column=(maxmonth + 2))
        #data_col = v1.cell(row=(row + len(office)), column=(node[i].month + 1))
        data_all = v1.cell(row=(row + len(office)), column=(maxmonth + 2))
        #reference = reference + str(row + len(office))
        #condition = condition + str(row + len(office))
    elif tabletype == 2:
        grandtotal = v1.cell(row=(row + len(person)), column=1)
        #data_main = v1.cell(row=(row + person.index(node[i].agent)), column=(node[i].month + 1))
        #data_row = v1.cell(row=(row + person.index(node[i].agent)), column=(maxmonth + 2))
        #data_col = v1.cell(row=(row + len(person)), column=(node[i].month + 1))
        data_all = v1.cell(row=(row + len(person)), column=(maxmonth + 2))
        #reference = reference + str(row + len(person))
        #condition = condition + str(row + len(person))
    elif tabletype == 3:
        grandtotal = v1.cell(row=(row + len(person)), column=1)
        #data_main = v1.cell(row=(row + person.index(node[i].agent)), column=(office.index(node[i].location) + 2))
        #data_row = v1.cell(row=(row + person.index(node[i].agent)), column=(maxmonth + 2))
        #data_col = v1.cell(row=(row + len(person)), column=(office.index(node[i].location) + 2))
        data_all = v1.cell(row=(row + len(person)), column=(maxmonth + 2))
        #reference = reference + str(row + len(person))
        #condition = condition + str(row + len(person))

    grandtotal.value = "Grand Total"
    grandtotal.font = openpyxl.styles.Font(bold = True)
    grandtotal.alignment = openpyxl.styles.Alignment(horizontal='center')
    grandtotal.fill = titlefill
    for n in node:
        if tabletype == 1:
            data_main = v1.cell(row=(row + office.index(node[i].location)), column=(node[i].month + 1))
            data_row = v1.cell(row=(row + office.index(node[i].location)), column=(maxmonth + 2))
            data_col = v1.cell(row=(row + len(office)), column=(node[i].month + 1))
        elif tabletype == 2:
            data_main = v1.cell(row=(row + person.index(node[i].agent)), column=(node[i].month + 1))
            data_row = v1.cell(row=(row + person.index(node[i].agent)), column=(maxmonth + 2))
            data_col = v1.cell(row=(row + len(person)), column=(node[i].month + 1))
        elif tabletype == 3:
            data_main = v1.cell(row=(row + person.index(node[i].agent)), column=(office.index(node[i].location) + 2))
            data_row = v1.cell(row=(row + person.index(node[i].agent)), column=(maxmonth + 2))
            data_col = v1.cell(row=(row + len(person)), column=(office.index(node[i].location) + 2))

        try:
            #When all data has no problem
            data_main.value += node[i].vol      #AddMain data
            data_row.value += node[i].vol       #Total of row
            data_col.value += node[i].vol       #Total of column
            data_all.value += node[i].vol       #Total of all
            pass
        except:
            #when data is Null
            if isinstance(data_main.value, int):        #Add Main data
                data_main.value += node[i].vol    
            else:
                data_main.value = node[i].vol
            data_main.alignment = openpyxl.styles.Alignment(horizontal='center')
                    
            if isinstance(data_col.value, int):         #Total of column
                data_col.value += node[i].vol                       
            else:
                data_col.value = node[i].vol
            data_col.font = openpyxl.styles.Font(bold = True)
            data_col.alignment = openpyxl.styles.Alignment(horizontal='center')
            data_col.fill = titlefill
                
            if isinstance(data_all.value, int):         #Total of all
                data_all.value += node[i].vol                            
            else:
                data_all.value = node[i].vol
            data_all.font = openpyxl.styles.Font(bold = True)
            data_all.alignment = openpyxl.styles.Alignment(horizontal='center')
            data_all.fill = titlefill
                
            if isinstance(data_row.value, int):         #Total of row
                data_row.value += node[i].vol         
            else:
                data_row.value = node[i].vol
            data_row.font = openpyxl.styles.Font(bold = True)
            data_row.alignment = openpyxl.styles.Alignment(horizontal='center')
            pass
        i += 1
    #v1.auto_filter.ref = reference
    #v1.auto_filter.add_sort_condition(condition)

def profitset(v2, node, row, office, person, maxmonth, tabletype):
    #tabletype= 1: By agent, 2: By office, 3: By month
    i = 0
    if tabletype == 1:
        grandtotal = v2.cell(row=(row + len(office)), column=1)
        data_all = v2.cell(row=(row + len(office)), column=(maxmonth + 2))
    elif tabletype == 2:
        grandtotal = v2.cell(row=(row + len(person)), column=1)
        data_all = v2.cell(row=(row + len(person)), column=(maxmonth + 2))
    elif tabletype == 3:
        grandtotal = v2.cell(row=(row + len(person)), column=1)
        data_all = v2.cell(row=(row + len(person)), column=(maxmonth + 2))

    grandtotal.value = "Grand Total"
    grandtotal.font = openpyxl.styles.Font(bold = True)
    grandtotal.alignment = openpyxl.styles.Alignment(horizontal='center')
    grandtotal.fill = titlefill
    for n in node:
        #In case that profit value is empty
        if node[i].profit is None:
            break
        #dataobject set for prepare multiple options
        if tabletype == 1:
            data_main = v2.cell(row=(row + office.index(node[i].location)), column=(node[i].month + 1))
            data_row = v2.cell(row=(row + office.index(node[i].location)), column=(maxmonth + 2))
            data_col = v2.cell(row=(row + len(office)), column=(node[i].month + 1))
        elif tabletype == 2:
            data_main = v2.cell(row=(row + person.index(node[i].agent)), column=(node[i].month + 1))
            data_row = v2.cell(row=(row + person.index(node[i].agent)), column=(maxmonth + 2))
            data_col = v2.cell(row=(row + len(person)), column=(node[i].month + 1))
        elif tabletype == 3:
            data_main = v2.cell(row=(row + person.index(node[i].agent)), column=(office.index(node[i].location) + 2))
            data_row = v2.cell(row=(row + person.index(node[i].agent)), column=(maxmonth + 2))
            data_col = v2.cell(row=(row + len(person)), column=(office.index(node[i].location) + 2))

        try:
            #When all data has no problem
            data_main.value += node[i].profit       #AddMain data
            data_row.value += node[i].profit        #Total of row
            data_col.value += node[i].profit        #Total of column
            data_all.value += node[i].profit        #Total of all
            pass
        except:
            #when data is Null
            if isinstance(data_main.value, float):  #Add Main data
                data_main.value += node[i].profit    
            else:
                #if type(node[i].profit) == 
                data_main.value = node[i].profit
            data_main.alignment = openpyxl.styles.Alignment(horizontal='center')
                    
            if isinstance(data_col.value, float):   #Total of column
                data_col.value += node[i].profit                       
            else:
                data_col.value = node[i].profit
            data_col.font = openpyxl.styles.Font(bold = True)
            data_col.alignment = openpyxl.styles.Alignment(horizontal='center')
            data_col.fill = titlefill
                
            if isinstance(data_all.value, float):   #Total of all
                data_all.value += node[i].profit                            
            else:
                data_all.value = node[i].profit
            data_all.font = openpyxl.styles.Font(bold = True)
            data_all.alignment = openpyxl.styles.Alignment(horizontal='center')
            data_all.fill = titlefill
                
            if isinstance(data_row.value, float):   #Total of row
                data_row.value += node[i].profit
            else:
                data_row.value = node[i].profit
            data_row.font = openpyxl.styles.Font(bold = True)
            data_row.alignment = openpyxl.styles.Alignment(horizontal='center')
            pass
        i += 1

def columnsize(v):
    y = 1
    max_col_size = 0
    for cols in v.columns:         #columns
        x = 1
        for rows in v.rows:        #rows
            if v.cell(row=x, column=y).value is None:
                pass
            else:
                if max_col_size < len(str(v.cell(row=x, column=y).value)):
                    max_col_size = len(str(v.cell(row=x, column=y).value))
            x += 1
        if y == 1:
            v.column_dimensions["A"].width = max_col_size
            max_col_size = 0
        y += 1
    y = 1
    for cols in v.columns:
        if y > 1:
            v.column_dimensions[openpyxl.utils.get_column_letter(y)].width = max_col_size
        y += 1


def agentlist(v3, person):
    if person is None:
        print("Wrong input function - \'Agent List\'")
    else:
        i = 0
        v3.cell(row=(i+1), column=1).value = "Agent List"
        v3.cell(row=(i+1), column=1).fill = titlefill
        while (i+1) <= (len(person)):
            v3.cell(row=(i+2), column=1).value = person[i]
            v3.cell(row=(i+2), column=1).fill = datafill
            i += 1

#==================== Main ====================
#load excel
workbook = openpyxl.load_workbook(filename="raw.xlsx")
sheet = workbook["Raw Data"]

#Build new sheet for store report data
workbook2 = openpyxl.Workbook()
v1 = workbook2.active
v1.title = "Volume"

v2 = workbook2.create_sheet()
v2.title = "Profit"

v3 = workbook2.create_sheet()
v3.title = "Agent List"

#set data
i = 1
maxmonth = 1
monthval = ""
node = []           #data store
office = []         #for by agent table
person = []         #for by office table
month = 9           #it will define with drop down
filter_node = []    #for by month table
filter_person = []  #for by month table

#==================== GUI get Month ====================
print("GUI....")
months = []
x = 1
while (x <= maxmonth):
    months.append(x)
    x += 1
    
def callbackFunc(event):
    app.quit()
app = tk.Tk() 
app.geometry('200x100')
app.title("Volume Report Generator")

labelTop = tk.Label(app, text = "Choose the month for Volumn Report")
labelTop.grid(column=0, row=0)

comboExample = ttk.Combobox(app, values=["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"], state="readonly")
comboExample.grid(column=0, row=1)
comboExample.current(0)
comboExample.bind("<<ComboboxSelected>>", callbackFunc)

app.mainloop()

#comboExample.current(): index of month(need to +1)
#comboExample.get(): value of month
month = (comboExample.current()+1)
monthval = comboExample.get()
print("Month: ", monthval, " / Value: ", month)

#==================== data input into class ====================
#find start point number of row
print("Define title organization....", end=" ")
row = 1
while (True):
    if sheet.cell(row=row, column=1).value is None:
        row += 1
    else:
        check = sheet.cell(row=row, column=1).value
        #get row of rawdata title
        if(check.lower() == "month" or check.lower() == "office" or check.lower() == "agent" or check.lower() == "volume" or check.lower() == "profit"):
            break
        row += 1
        #if(isinstance(check, int)):
        #    break
print("Done")

print("Gething data from raw data....")
a = b = c = d = e = col = 1
while (True):
    check = sheet.cell(row=row, column=col).value
    if sheet.cell(row=row, column=col).value is None:
        break
    else:
        # Month     : a     Office    : b       Agent     : c       Volume    : d       Profit    : e
        if "month" in check.lower():
            print("Pass:\tCol: ", col, " : Month")
            a = col
        elif "office" in check.lower():
            print("Pass:\tCol: ", col, " : Office")
            b = col
        elif "agent" in check.lower():
            print("Pass:\tCol: ", col, " : Agent")
            c = col
        elif "volume" in check.lower():
            print("Pass:\tCol: ", col, " : Volume")
            d = col
        elif "profit" in check.lower():
            print("Pass:\tCol: ", col, " : Profit")
            e = col
    col += 1
print("Done")

print("Set gathered data....", end=" ")
row += 1
while (True):
    if sheet.cell(row=row, column=1).value is None:
        break
    else:
        node.append(Node(sheet.cell(row=row, column=a).value, sheet.cell(row=row, column=b).value, sheet.cell(row=row, column=c).value, sheet.cell(row=row, column=d).value, sheet.cell(row=row, column=e).value))
        if maxmonth < sheet.cell(row=row, column=a).value:
            maxmonth = sheet.cell(row=row, column=a).value
        #get list of office
        if sheet.cell(row=row, column=b).value not in office:
            office.append(sheet.cell(row=row, column=b).value)
        #get list of agent
        if sheet.cell(row=row, column=c).value not in person:
            person.append(sheet.cell(row=row, column=c).value)
        #get filtered (monthly) person node
        if sheet.cell(row=row, column=a).value == month:
            filter_node.append(Node(sheet.cell(row=row, column=a).value, sheet.cell(row=row, column=b).value, sheet.cell(row=row, column=c).value, sheet.cell(row=row, column=d).value, sheet.cell(row=row, column=e).value))
            if sheet.cell(row=row, column=c).value not in filter_person:
                filter_person.append(sheet.cell(row=row, column=c).value)
    row += 1
print("Done")

#==================== Output - Volume ====================
row = 2

#By Agent
print("Volume: By Agent....", end=" ")
row = title(v1, row, maxmonth, office, 1, None)
label(v1, row, office, person, 1)
dataset(v1, node, row, office, person, maxmonth, 1)
row = row + len(office) + 4
print("Done")

#By Office
print("Volume: By Office....", end=" ")
row = title(v1, row, maxmonth, office, 2, None)
label(v1, row, office, person, 2)
dataset(v1, node, row, office, person, maxmonth, 2)
row = row + len(person) + 4
print("Done")

#By Month
print("Volume: By Month....", end=" ")
row = title(v1, row, len(office), office, 3, monthval)
label(v1, row, office, filter_person, 3)
dataset(v1, filter_node, row, office, filter_person, len(office), 3)
print("Done")

print("Volume: Design....", end=" ")
columnsize(v1)
print("Done")

#==================== Output - Profit ====================
row = 2

#By Agent
print("Profit: By Agent....", end=" ")
row = title(v2, row, maxmonth, office, 1, None)
label(v2, row, office, person, 1)
profitset(v2, node, row, office, person, maxmonth, 1)
row = row + len(office) + 4
print("Done")

#By Office
print("Profit: By Office....", end=" ")
row = title(v2, row, maxmonth, office, 2, None)
label(v2, row, office, person, 2)
profitset(v2, node, row, office, person, maxmonth, 2)
row = row + len(person) + 4
print("Done")

#By Month
print("Profit: By Month....", end=" ")
row = title(v2, row, len(office), office, 3, monthval)
label(v2, row, office, filter_person, 3)
profitset(v2, filter_node, row, office, filter_person, len(office), 3)
print("Done")

print("Profit: Design....", end=" ")
columnsize(v2)
print("Done")

#==================== Output - Agent List ==================== 
print("AgentList: Display....", end=" ")
agentlist(v3, person)
print("Done")

print("AgentList: Design....", end=" ")
columnsize(v3)
print("Done")

print("Save as Excel....", end=" ")
workbook2.save(filename="TPEB FAK Volume and Profit Report (All Binex).xlsx")
print("Done")